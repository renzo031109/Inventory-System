from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from inventory.models import Item, ItemBase
from django.http import HttpResponse
import datetime

from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
from openpyxl.styles import *

#set universal variable for value critical
critical_value = 10

# Create your views here.
@login_required
def dashboard_view(request):

    #initialize critical value
    
    itembase = ItemBase.objects.all()
    item = Item.objects.all()

    #Total Item Count
    item_count = itembase.count()

    #Total Items with critical stocks
    critical_count = 0
    for critical in itembase:
        if critical.soh <= critical_value:
            critical_count += 1

    #Transaction today
    transaction_count = 0
    for transaction_date in item:
        if transaction_date.date_added.date() == datetime.datetime.now().date():
            transaction_count += 1
            

    context = {
        'item_count': item_count,
        'critical_count': critical_count,
        'itembase': itembase,
        'transaction_count': transaction_count

    }

    return render(request, 'dashboard/dashboard_template.html', context)




def critical_stock_excel_export(request):
    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ITEM WITH CRITICAL STOCK.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:G1')

    first_cell = worksheet['A1']
    first_cell.value = "ITEM WITH CRITICAL STOCK"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "ITEM WITH CRITICAL STOCK"

    # Add headers
    headers =   [
                'ITEM NAME',	
                'BRAND NAME',
                'UOM',	
                'SOH',	
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    items = ItemBase.objects.all()


    for item in items:
        
        if item.soh <= critical_value :
            #convert object fields to string
            uom = str(item.uom)

            worksheet.append([
                item.item_name,
                item.brand_name,
                uom,
                item.soh,
            ])
    
    workbook.save(response)
    return response



