from typing import Any
from django.forms import BaseModelForm
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from .models import Item, ItemBase, ItemCode, UOM, Client, Department
from .forms import ItemNewForm, ItemModelFormSet
from .filters import ItemFilter, ItemBaseFilter
from django.http import HttpResponse
# from django.core.paginator import Paginator

#for export excel imports
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import *


#set universal variable for user settings.
user_department = "FINANCE"
user_client = "SHORE360"

@login_required
def inventory_item(request):
    
    items = Item.objects.all()
    # item_count_total= items.count()
    
    itemFilter = ItemFilter(request.GET, queryset=items)
    items = itemFilter.qs
    
    item_count = items.count()
    
    if item_count > 0 :
        messages.info(request, f"Found '{item_count}' transaction in the database")
    else:
        messages.info(request, f"Item not Found in the database ")
    
    # #pagination show 50 items per page
    # paginator = Paginator(items, 25)
    # page_number = request.GET.get("page")
    # page_obj = paginator.get_page(page_number)

    context = {
        'items': items, 
        'item_count': item_count, 
        'itemFilter': itemFilter,
        # 'page_obj': page_obj
        }
    return render(request,'inventory/inventory.html', context)


@login_required
def delete_item(request, id):
    if request.method == 'POST':
        #get the selected value
        item = Item.objects.get(id=id)
        item_soh = ItemBase.objects.get(item_code=item.item_code)

        #return the quantity of the deleted item
        updated_soh = int(item_soh.soh) + int(item.quantity)
        item_soh.soh = updated_soh

        #update Total value and save
        total = updated_soh * item_soh.price
        item_soh.total_value = total
        item_soh.save()
        item.delete()
        messages.success(request, "Item is deleted successfully")
    return redirect('inventory_item')



def delete_itembase(request, item_code):
    if request.method == 'POST':
        item = ItemBase.objects.get(item_code=item_code)
        itemcode = ItemCode.objects.get(code=item)

        print(item)
        print(itemcode)

        item.delete()
        itemcode.delete()

        messages.success(request, f"{item.item_name} is deleted successfully")
    return redirect('summary_item')



@login_required
def summary_item(request):
    items = ItemBase.objects.all()
    # item_count_total= items.count()

    itemFilter = ItemBaseFilter(request.GET, queryset=items)
    items = itemFilter.qs
    item_count = items.count()

    if item_count > 0 :
        messages.info(request, f"Found '{item_count}' item(s) in the database")
    else:
        messages.info(request, f"Item not Found in the database ")

    # #pagination show 50 items per page
    # paginator = Paginator(items, 25)
    # page_number = request.GET.get("page")
    # page_obj = paginator.get_page(page_number)

    context = {
        'items': items,
        'item_count': item_count,
        'itemFilter': itemFilter,
        # 'page_obj':page_obj
        }
    return render(request, 'inventory/summary.html', context)


@login_required
def new_item(request):
    #initialize user client and department static

    if request.method == 'POST':
        form = ItemNewForm(request.POST)

        if form.is_valid():
            #get the value of the form
            form_item_name = request.POST.get('item_name')
            form_item_brand = request.POST.get('brand_name')
            form_item_soh = request.POST.get('soh')
            form_item_uom = request.POST.get('uom')
            form_item_price = request.POST.get('price')

            #convert UOM id to values of foreign key
            uom_value = UOM.objects.get(id=form_item_uom)
         
            #using try-except method in case of null value
            try:
                record_name = ItemBase.objects.filter(item_name=form_item_name, brand_name=form_item_brand)

                for record in record_name:
                    if record.item_name.upper() == form_item_name.upper() and record.brand_name.upper() == form_item_brand.upper():
                        messages.error(request, "Item already exist!")
                        return redirect('new_item')
                       
            except:
                record_name = None

            #assign default value to remarks
            concat = form_item_name + " | " + form_item_brand
            itemcode = ItemCode(code=concat)
            itemcode.save()

            #Assign form to a variable
            itemAddForm = form.save(commit=False)

            #define user for the staffname
            user = request.user
            client = Client.objects.get(client=user_client)
            department=Department.objects.get(department=user_department)

            #copy newitem to Item Transaction
            itemTransaction = Item(
                                    item_code=itemcode, 
                                    item_name=form_item_name, 
                                    brand_name=form_item_brand, 
                                    quantity=form_item_soh, 
                                    remarks="BEGINNING",
                                    uom=uom_value, 
                                    staff_name=user.username,
                                    client_name=client,
                                    department_name=department
                                    )
            itemTransaction.save()

            #assign generated code value to itemcode 
            itemAddForm.item_code = concat
            itemAddForm.total_value = int(form_item_price) * int(form_item_soh)
            itemAddForm.save()

            messages.success(request, "New Item added successfully!")
            return redirect('summary_item')
    else:
        form = ItemNewForm()

    context = {'form': form}
    return render(request, 'inventory/new_item.html', context)


@login_required
def add_item(request):

    if request.method == 'POST':
        formset = ItemModelFormSet(request.POST)
        if formset.is_valid():
            for form in formset:
                
                # check if itemcode is selected
                if form.cleaned_data.get('item_code'):  
                    
                    #get the item name from the form 
                    add_item_code = form.cleaned_data.get('item_code')
                    #get the item qty from the form
                    add_qty = form.cleaned_data.get('quantity')
                    
                    
                    try:
                        #get the item SOH from model table
                        item_soh = ItemBase.objects.get(item_code=add_item_code)

                        #assign value to client and department of authenticated user
                        client = Client.objects.get(client=user_client)
                        department = Department.objects.get(department=user_department)
    
                        #compute add soh
                        soh = int(item_soh.soh) + int(add_qty)
                        #get the updated soh after add
                        item_soh.soh = int(soh)

                        #update Total value
                        total = item_soh.soh * item_soh.price
                        item_soh.total_value = total

                        #save tables
                        item_soh.save()
                    except:
                        item_soh = 0
                        soh = int(item_soh) + int(add_qty)

                    #assign default values  
                    itemAddForm = form.save(commit=False)    
                    itemAddForm.remarks = "IN"
                    itemAddForm.item_name = item_soh.item_name
                    itemAddForm.brand_name = item_soh.brand_name
                    itemAddForm.uom = item_soh.uom

                    #get the current user
                    user = request.user
                    itemAddForm.staff_name = user.username
                    itemAddForm.client_name = client
                    itemAddForm.department_name = department
  
                    itemAddForm.save()
                    
            messages.success(request, "You added stock successfully!")
            
            return redirect('inventory_item')
        else:
            messages.error(request, "Invalid Input!")

    else:
        formset = ItemModelFormSet(queryset=Item.objects.none())

    context = {'formset': formset}
    return render(request, 'inventory/add_item.html', context)


@login_required
def get_item(request):

    #Initiate a list variable for the input select fields
    staff_name_list = []
    client_name_list = []
    department_name_list = []

    #list for validation checking
    item_error_list = []
    item_success_list = []

    if request.method == 'POST':
        formset = ItemModelFormSet(request.POST)
        if formset.is_valid():
            for form in formset:

                # only save if name is present
                print(form.cleaned_data.get('item_code'))
                print(form.cleaned_data.get('quantity'))
                if form.cleaned_data.get('item_code'):  
                    
                    #get the item name from the form 
                    get_item_code = form.cleaned_data.get('item_code')

                    #get the item qty from the form
                    get_qty = form.cleaned_data.get('quantity')

                    #get the item SOH from model table       
                    item_soh = ItemBase.objects.get(item_code=get_item_code)

                    #get the staff name values
                    get_staff_name = form.cleaned_data.get('staff_name')
                    get_client_name = form.cleaned_data.get('client_name')
                    get_department_name = form.cleaned_data.get('department_name')

                    #populate the list from the user input
                    staff_name_list.append(get_staff_name)
                    client_name_list.append(get_client_name)
                    department_name_list.append(get_department_name)
                    
                    #get the first value of the form
                    staff_name = staff_name_list[0]
                    client_name = client_name_list[0]
                    department_name = department_name_list[0]

                    if item_soh.soh < get_qty:
                        messages.error(request, f"Ooops, Your available stock for '{item_soh.item_name}' is only '{item_soh.soh}")
                        item_error_list.append(item_soh.item_name)
                    else:
                        item_success_list.append(item_soh.item_name)
                        #compute add soh
                        soh = int(item_soh.soh) - int(get_qty)
                        #get the updated soh after add
                        item_soh.soh = int(soh)
                        #update Total value
                        total = item_soh.soh * item_soh.price
                        item_soh.total_value = total
                        #save tables
                        item_soh.save()

                        #assign default value to remarks 
                        itemGetForm = form.save(commit=False)    
                        itemGetForm.remarks = "OUT"
                        itemGetForm.item_name = item_soh.item_name
                        itemGetForm.brand_name = item_soh.brand_name
                        itemGetForm.uom = item_soh.uom
                        itemGetForm.staff_name = staff_name
                        itemGetForm.client_name = client_name
                        itemGetForm.department_name = department_name
                        itemGetForm.save()
                        
            #assign length variables
            invalid_form = len(item_error_list)
            valid_form = len(item_success_list)

            # if values are valid send to submitted template
            if invalid_form == 0 and valid_form != 0:
                return redirect('submitted')

            # if 1 of the values are valid 
            elif invalid_form > 0 and valid_form > 0:
                messages.info(request, f"{valid_form} item(s) are submitted. Please re-input the item(s) with insufficient stock on hand.")
                return redirect('get_item')    
            # re input the items     
            else:
                return redirect('get_item')
            
        else:
            messages.error(request, "Invalid Input!")

    else:
        formset = ItemModelFormSet(queryset=Item.objects.none())

    context = {'formset': formset}
    return render(request, 'inventory/get_item.html', context)


@login_required
def submitted(request):
    return render(request, 'inventory/submitted.html')


@login_required
def export_excel_inventory(request):

    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ITEM INVENTORY REPORT.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:I1')

    first_cell = worksheet['A1']
    first_cell.value = "ITEM INVENTORY REPORT"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "Inventory Report"

    # Add headers
    headers =   [
                'ITEM NAME',	
                'BRAND NAME',
                'QUANTITY',	
                'UOM',	
                'DATE ADDED',
                'REMARKS',	
                'STAFF NAME',
                'CLIENT NAME',
                'DEPARTMENT NAME'	
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    items = Item.objects.all()
    for item in items:
        
        #convert object fields to string
        client_name = str(item.client_name)
        department_name = str(item.department_name)
        date_added = datetime.strftime(item.date_added,'%m/%d/%Y %H:%M:%S')

        worksheet.append([
            item.item_name,
            item.brand_name,
            item.quantity,
            item.uom,
            date_added,
            item.remarks,
            item.staff_name,
            client_name,
            department_name
        ])
    
    workbook.save(response)
    return response


@login_required
def export_excel_summary(request):

    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="SUMMARY ITEM REPORT.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:G1')

    first_cell = worksheet['A1']
    first_cell.value = "SUMMARY ITEM REPORT"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "SUMMARY ITEM REPORT"

    # Add headers
    headers =   [
                'ITEM NAME',	
                'BRAND NAME',
                'UOM',	
                'SOH',
                'PRICE',	
                'TOTAL VALUE',
                'DATE ADDED'
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    items = ItemBase.objects.all()
    for item in items:
        
        #convert object fields to string

        uom = str(item.uom)
        date_added = datetime.strftime(item.date_added,'%m/%d/%Y %H:%M:%S')

        worksheet.append([
            item.item_name,
            item.brand_name,
            uom,
            item.soh,
            item.price,
            item.total_value,
            date_added,
        ])
    
    workbook.save(response)
    return response



